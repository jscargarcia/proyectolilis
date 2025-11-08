from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.db import transaction
from decimal import Decimal, InvalidOperation
from autenticacion.decorators import login_required_custom, permission_required, estado_usuario_activo, permiso_requerido
# Importar modelos desde maestros donde están definidos
from .models import Producto, Proveedor, Categoria, Marca, UnidadMedida
# Para exportación a Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
# Para manejo de imágenes
from PIL import Image
import os
import re


# ==================== PRODUCTOS ====================

@login_required_custom
@estado_usuario_activo
def producto_listar(request):
    """Lista de productos con búsqueda, filtros, paginación y ordenamiento"""
    # Parámetros de búsqueda y filtros
    query = request.GET.get('query', '').strip()
    categoria_id = request.GET.get('categoria', '')
    marca_id = request.GET.get('marca', '')
    estado = request.GET.get('estado', '')
    
    # Parámetros de ordenamiento y paginación con mantenimiento en sesión
    orden = request.GET.get('orden', 'nombre')
    items_per_page = request.GET.get('items_per_page')
    
    # Si no se especifica en URL, usar valor de sesión o default
    if not items_per_page:
        items_per_page = request.session.get('productos_items_per_page', '15')
    
    # Validar items por página con nuevas opciones
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [5, 15, 30, 50]:
            items_per_page = 15
    except (ValueError, TypeError):
        items_per_page = 15
    
    # Guardar selección en sesión
    request.session['productos_items_per_page'] = str(items_per_page)
    
    # Query base con relaciones mejoradas
    productos = Producto.objects.select_related('categoria', 'marca', 'uom_compra', 'uom_venta', 'uom_stock').all()
    
    # Aplicar filtros mejorados con búsqueda por precio y stock
    if query:
        # Intentar convertir query a número para búsqueda por precio/stock
        try:
            query_num = Decimal(query)
            productos = productos.filter(
                Q(sku__icontains=query) |
                Q(nombre__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(categoria__nombre__icontains=query) |
                Q(marca__nombre__icontains=query) |
                Q(modelo__icontains=query) |
                Q(precio_venta=query_num) |
                Q(stock_minimo=query_num) |
                Q(ean_upc__icontains=query)
            )
        except (ValueError, InvalidOperation):
            # Si no es número, buscar solo en campos de texto
            productos = productos.filter(
                Q(sku__icontains=query) |
                Q(nombre__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(categoria__nombre__icontains=query) |
                Q(marca__nombre__icontains=query) |
                Q(modelo__icontains=query) |
                Q(ean_upc__icontains=query)
            )
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    
    if estado:
        productos = productos.filter(estado=estado)
    
    # Aplicar ordenamiento
    orden_mapping = {
        'nombre': 'nombre',
        'nombre_desc': '-nombre',
        'sku': 'sku',
        'sku_desc': '-sku',
        'categoria': 'categoria__nombre',
        'categoria_desc': '-categoria__nombre',
        'marca': 'marca__nombre',
        'marca_desc': '-marca__nombre',
        'precio': 'precio_venta',
        'precio_desc': '-precio_venta',
        'stock': 'stock_minimo',
        'stock_desc': '-stock_minimo',
        'fecha': 'created_at',
        'fecha_desc': '-created_at',
    }
    
    if orden in orden_mapping:
        productos = productos.order_by(orden_mapping[orden])
    else:
        productos = productos.order_by('nombre')
    
    # Paginación
    paginator = Paginator(productos, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para filtros
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    
    # Estadísticas
    total_productos = productos.count()
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'categoria_id': categoria_id,
        'marca_id': marca_id,
        'estado': estado,
        'orden': orden,
        'items_per_page': items_per_page,
        'categorias': categorias,
        'marcas': marcas,
        'total_productos': total_productos,
        'items_per_page_options': [5, 15, 30, 50],
        'orden_options': [
            ('nombre', 'Nombre A-Z'),
            ('nombre_desc', 'Nombre Z-A'),
            ('sku', 'SKU A-Z'),
            ('sku_desc', 'SKU Z-A'),
            ('categoria', 'Categoría A-Z'),
            ('categoria_desc', 'Categoría Z-A'),
            ('marca', 'Marca A-Z'),
            ('marca_desc', 'Marca Z-A'),
            ('precio', 'Precio Menor'),
            ('precio_desc', 'Precio Mayor'),
            ('stock', 'Stock Menor'),
            ('stock_desc', 'Stock Mayor'),
            ('fecha', 'Más Antiguos'),
            ('fecha_desc', 'Más Recientes'),
        ]
    }
    return render(request, 'maestros/producto_listar.html', context)


@login_required_custom
@estado_usuario_activo
@permiso_requerido('productos', 'crear')
def producto_crear(request):
    """Crear nuevo producto con validaciones simplificadas para el modelo productos"""
    if request.method == 'POST':
        errores = []
        
        # Validar campos requeridos según el modelo productos
        sku = request.POST.get('sku', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        categoria_id = request.POST.get('categoria', '')
        marca_id = request.POST.get('marca', '')
        
        if not sku:
            errores.append('El SKU es requerido.')
        elif len(sku) < 3:
            errores.append('El SKU debe tener al menos 3 caracteres.')
        elif Producto.objects.filter(sku=sku).exists():
            errores.append('Ya existe un producto con este SKU.')
            
        if not nombre:
            errores.append('El nombre es requerido.')
        elif len(nombre) < 3:
            errores.append('El nombre debe tener al menos 3 caracteres.')
            
        if not categoria_id:
            errores.append('La categoría es requerida.')
        elif not Categoria.objects.filter(id=categoria_id, activo=True).exists():
            errores.append('La categoría seleccionada no es válida.')
        
        # Validar precios y stock con validaciones mejoradas
        precio_venta = request.POST.get('precio_venta', '').strip()
        costo_estandar = request.POST.get('costo_estandar', '').strip()
        stock_minimo = request.POST.get('stock_minimo', '0').strip()
        stock_maximo = request.POST.get('stock_maximo', '').strip()
        ean_upc = request.POST.get('ean_upc', '').strip()
        
        # Validar precio de venta (debe ser > 0)
        if precio_venta:
            try:
                precio_venta = Decimal(precio_venta)
                if precio_venta <= 0:
                    errores.append('El precio de venta debe ser mayor a 0.')
            except (ValueError, InvalidOperation):
                errores.append('El precio de venta debe ser un número válido.')
                precio_venta = None
        else:
            precio_venta = None
        
        # Validar costo estándar (debe ser >= 0)
        if costo_estandar:
            try:
                costo_estandar = Decimal(costo_estandar)
                if costo_estandar < 0:
                    errores.append('El costo estándar no puede ser negativo.')
            except (ValueError, InvalidOperation):
                errores.append('El costo estándar debe ser un número válido.')
                costo_estandar = None
        else:
            costo_estandar = None
        
        # Validar stock mínimo (debe ser >= 0)    
        try:
            stock_minimo = Decimal(stock_minimo) if stock_minimo else 0
            if stock_minimo < 0:
                errores.append('El stock mínimo no puede ser negativo.')
        except (ValueError, TypeError, InvalidOperation):
            errores.append('El stock mínimo debe ser un número válido.')
            stock_minimo = 0
        
        # Validar stock máximo (debe ser >= stock mínimo)
        if stock_maximo:
            try:
                stock_maximo = Decimal(stock_maximo)
                if stock_maximo < 0:
                    errores.append('El stock máximo no puede ser negativo.')
                elif stock_maximo < stock_minimo:
                    errores.append('El stock máximo debe ser mayor o igual al stock mínimo.')
            except (ValueError, InvalidOperation):
                errores.append('El stock máximo debe ser un número válido.')
                stock_maximo = None
        else:
            stock_maximo = None
        
        # Validar código de barras (debe ser único si se proporciona)
        if ean_upc:
            if len(ean_upc) < 8 or len(ean_upc) > 20:
                errores.append('El código de barras debe tener entre 8 y 20 caracteres.')
            elif Producto.objects.filter(ean_upc=ean_upc).exists():
                errores.append('Ya existe un producto con este código de barras.')
        
        # Validar unidades de medida requeridas
        uom_compra_id = request.POST.get('uom_compra', '')
        uom_venta_id = request.POST.get('uom_venta', '')
        uom_stock_id = request.POST.get('uom_stock', '')
        
        if not uom_compra_id:
            errores.append('La unidad de medida de compra es requerida.')
        elif not UnidadMedida.objects.filter(id=uom_compra_id, activo=True).exists():
            errores.append('La unidad de medida de compra seleccionada no es válida.')
            
        if not uom_venta_id:
            errores.append('La unidad de medida de venta es requerida.')
        elif not UnidadMedida.objects.filter(id=uom_venta_id, activo=True).exists():
            errores.append('La unidad de medida de venta seleccionada no es válida.')
            
        if not uom_stock_id:
            errores.append('La unidad de medida de stock es requerida.')
        elif not UnidadMedida.objects.filter(id=uom_stock_id, activo=True).exists():
            errores.append('La unidad de medida de stock seleccionada no es válida.')
        
        # Validar factor de conversión
        factor_conversion = request.POST.get('factor_conversion', '1').strip()
        try:
            factor_conversion = Decimal(factor_conversion) if factor_conversion else 1
            if factor_conversion <= 0:
                errores.append('El factor de conversión debe ser mayor a 0.')
        except (ValueError, InvalidOperation):
            errores.append('El factor de conversión debe ser un número válido.')
            factor_conversion = 1
        
        # Validar imagen subida
        imagen = request.FILES.get('imagen')
        if imagen:
            # Validar tipo de archivo
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
            if imagen.content_type not in allowed_types:
                errores.append('La imagen debe ser JPG, PNG o GIF.')
            
            # Validar tamaño (máximo 5MB)
            if imagen.size > 5 * 1024 * 1024:  # 5MB en bytes
                errores.append('La imagen no puede superar los 5MB.')
            
            # Validar dimensiones usando PIL
            try:
                img = Image.open(imagen)
                width, height = img.size
                
                # Recomendaciones de dimensiones
                if width > 2000 or height > 2000:
                    errores.append('Se recomienda que la imagen no supere 2000x2000 píxeles.')
                
                img.close()
            except Exception as e:
                errores.append('El archivo de imagen no es válido.')
        
        imagen_url = request.POST.get('imagen_url', '').strip()
        
        # Si hay errores, mostrarlos
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            # Crear producto con todos los campos validados
            try:
                with transaction.atomic():
                    producto = Producto.objects.create(
                        sku=sku,
                        ean_upc=ean_upc if ean_upc else None,
                        nombre=nombre,
                        descripcion=request.POST.get('descripcion', '').strip(),
                        categoria_id=categoria_id,
                        marca_id=marca_id if marca_id else None,
                        modelo=request.POST.get('modelo', '').strip(),
                        uom_compra_id=uom_compra_id,
                        uom_venta_id=uom_venta_id,
                        uom_stock_id=uom_stock_id,
                        factor_conversion=factor_conversion,
                        costo_estandar=costo_estandar,
                        precio_venta=precio_venta,
                        impuesto_iva=Decimal(request.POST.get('impuesto_iva', '19')),
                        stock_minimo=stock_minimo,
                        stock_maximo=stock_maximo,
                        punto_reorden=Decimal(request.POST.get('punto_reorden', '0')) if request.POST.get('punto_reorden') else None,
                        perishable=request.POST.get('perishable') == 'on',
                        control_por_lote=request.POST.get('control_por_lote') == 'on',
                        control_por_serie=request.POST.get('control_por_serie') == 'on',
                        imagen=imagen if imagen else None,
                        imagen_url=imagen_url if imagen_url else None,
                        estado=request.POST.get('estado', 'ACTIVO'),
                    )
                    messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
                    return JsonResponse({
                        'success': True,
                        'message': f'Producto "{producto.nombre}" creado exitosamente.',
                        'redirect_url': reverse('maestros:producto_listar')
                    })
            except Exception as e:
                messages.error(request, f'Error al crear producto: {str(e)}')
                return JsonResponse({
                    'success': False,
                    'message': f'Error al crear producto: {str(e)}'
                })
        
        # Si llegamos aquí, hay errores
        return JsonResponse({
            'success': False,
            'errors': errores
        })
    
    # GET request - mostrar formulario
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    unidades_medida = UnidadMedida.objects.filter(activo=True).order_by('tipo', 'nombre')
    
    context = {
        'categorias': categorias,
        'marcas': marcas,
        'unidades_medida': unidades_medida,
        'estados_producto': [
            ('Activo', 'Activo'),
            ('Inactivo', 'Inactivo'),
        ],
    }
    return render(request, 'maestros/producto_crear.html', context)


@login_required_custom
@estado_usuario_activo
def producto_detalle(request, pk):
    """Detalle de producto"""
    producto = get_object_or_404(Producto.objects.select_related('categoria', 'marca'), pk=pk)
    # Obtener proveedores relacionados
    try:
        from productos.models import ProductoProveedor
        proveedores = ProductoProveedor.objects.filter(producto=producto, activo=True).select_related('proveedor')
    except:
        proveedores = []
    
    context = {
        'producto': producto,
        'proveedores': proveedores,
    }
    return render(request, 'maestros/producto_detalle.html', context)


@login_required_custom
@estado_usuario_activo
@permiso_requerido('productos', 'editar')
def producto_editar(request, pk):
    """Editar producto - Versión simplificada para modelo productos"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        errores = []
        
        # Validar campos requeridos
        sku = request.POST.get('sku', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        categoria_id = request.POST.get('categoria', '')
        marca_id = request.POST.get('marca', '')
        
        if not sku:
            errores.append('El SKU es requerido.')
        elif len(sku) < 3:
            errores.append('El SKU debe tener al menos 3 caracteres.')
        elif Producto.objects.filter(sku=sku).exclude(pk=producto.pk).exists():
            errores.append('Ya existe otro producto con este SKU.')
            
        if not nombre:
            errores.append('El nombre es requerido.')
        elif len(nombre) < 3:
            errores.append('El nombre debe tener al menos 3 caracteres.')
            
        if not categoria_id:
            errores.append('La categoría es requerida.')
        elif not Categoria.objects.filter(id=categoria_id, activo=True).exists():
            errores.append('La categoría seleccionada no es válida.')
            
        # Validar marca (opcional)
        if marca_id and not Marca.objects.filter(id=marca_id, activo=True).exists():
            errores.append('La marca seleccionada no es válida.')
        
        # Validar precio de venta
        precio_venta = request.POST.get('precio_venta', '').strip()
        if not precio_venta:
            errores.append('El precio de venta es requerido.')
        else:
            try:
                precio_venta = Decimal(precio_venta)
                if precio_venta <= 0:
                    errores.append('El precio de venta debe ser mayor a 0.')
            except (ValueError, InvalidOperation):
                errores.append('El precio de venta debe ser un número válido.')
                precio_venta = None
            
        # Validar EAN/UPC (opcional)
        ean_upc = request.POST.get('ean_upc', '').strip()
        if ean_upc and not ean_upc.isdigit():
            errores.append('El código EAN/UPC solo debe contener números.')
        elif ean_upc and len(ean_upc) not in [8, 12, 13]:
            errores.append('El código EAN/UPC debe tener 8, 12 o 13 dígitos.')
        
        # Validar costo estándar (opcional)
        costo_estandar = request.POST.get('costo_estandar', '').strip()
        if costo_estandar:
            try:
                costo_estandar = Decimal(costo_estandar)
                if costo_estandar < 0:
                    errores.append('El costo estándar no puede ser negativo.')
            except (ValueError, InvalidOperation):
                errores.append('El costo estándar debe ser un número válido.')
                costo_estandar = None
        else:
            costo_estandar = None
            
        # Validar IVA
        impuesto_iva = request.POST.get('impuesto_iva', '19').strip()
        try:
            impuesto_iva = Decimal(impuesto_iva)
            if impuesto_iva < 0 or impuesto_iva > 100:
                errores.append('El IVA debe estar entre 0 y 100%.')
        except (ValueError, InvalidOperation):
            errores.append('El IVA debe ser un número válido.')
            impuesto_iva = Decimal('19.00')
        
        # Validar stock mínimo
        stock_minimo = request.POST.get('stock_minimo', '0').strip()
        try:
            stock_minimo = int(stock_minimo)
            if stock_minimo < 0:
                errores.append('El stock mínimo no puede ser negativo.')
        except (ValueError, TypeError):
            errores.append('El stock mínimo debe ser un número entero válido.')
            stock_minimo = 0
            
        # Validar stock máximo (opcional)
        stock_maximo = request.POST.get('stock_maximo', '').strip()
        if stock_maximo:
            try:
                stock_maximo = int(stock_maximo)
                if stock_maximo < 0:
                    errores.append('El stock máximo no puede ser negativo.')
                elif stock_maximo < stock_minimo:
                    errores.append('El stock máximo debe ser mayor o igual al stock mínimo.')
            except (ValueError, TypeError):
                errores.append('El stock máximo debe ser un número entero válido.')
                stock_maximo = None
        else:
            stock_maximo = None
            
        # Validar punto de reorden (opcional)
        punto_reorden = request.POST.get('punto_reorden', '').strip()
        if punto_reorden:
            try:
                punto_reorden = int(punto_reorden)
                if punto_reorden < 0:
                    errores.append('El punto de reorden no puede ser negativo.')
            except (ValueError, TypeError):
                errores.append('El punto de reorden debe ser un número entero válido.')
                punto_reorden = None
        else:
            punto_reorden = None
            
        # Validar factor de conversión
        factor_conversion = request.POST.get('factor_conversion', '1').strip()
        try:
            factor_conversion = Decimal(factor_conversion)
            if factor_conversion <= 0:
                errores.append('El factor de conversión debe ser mayor a 0.')
        except (ValueError, InvalidOperation):
            errores.append('El factor de conversión debe ser un número válido.')
            factor_conversion = Decimal('1.0000')
        
        # Si hay errores, mostrarlos
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            # Actualizar producto
            try:
                with transaction.atomic():
                    # Información básica
                    producto.sku = sku
                    producto.ean_upc = ean_upc or None
                    producto.nombre = nombre
                    producto.descripcion = request.POST.get('descripcion', '').strip() or None
                    producto.modelo = request.POST.get('modelo', '').strip() or None
                    
                    # Clasificación
                    producto.categoria_id = categoria_id
                    producto.marca_id = marca_id if marca_id else None
                    producto.estado = request.POST.get('estado', 'Activo')
                    
                    # Unidades de medida
                    uom_compra_id = request.POST.get('uom_compra')
                    uom_venta_id = request.POST.get('uom_venta')
                    uom_stock_id = request.POST.get('uom_stock')
                    producto.uom_compra_id = uom_compra_id if uom_compra_id else None
                    producto.uom_venta_id = uom_venta_id if uom_venta_id else None
                    producto.uom_stock_id = uom_stock_id if uom_stock_id else None
                    producto.factor_conversion = factor_conversion
                    
                    # Precios y costos
                    producto.costo_estandar = costo_estandar
                    producto.precio_venta = precio_venta
                    producto.impuesto_iva = impuesto_iva
                    
                    # Control de stock
                    producto.stock_minimo = stock_minimo
                    producto.stock_maximo = stock_maximo
                    producto.punto_reorden = punto_reorden
                    
                    # Características
                    producto.perishable = request.POST.get('perishable') == 'on'
                    producto.control_por_lote = request.POST.get('control_por_lote') == 'on'
                    producto.control_por_serie = request.POST.get('control_por_serie') == 'on'
                    
                    # Imagen
                    imagen_url = request.POST.get('imagen_url', '').strip()
                    producto.imagen_url = imagen_url if imagen_url else None
                    
                    producto.save()
                    
                    messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
                    return JsonResponse({
                        'success': True,
                        'message': f'Producto "{producto.nombre}" actualizado exitosamente. Estado: {producto.get_estado_display()}',
                        'redirect_url': reverse('maestros:producto_detalle', kwargs={'pk': producto.pk}),
                        'nuevo_estado': producto.estado,
                        'estado_display': producto.get_estado_display()
                    })
            except Exception as e:
                messages.error(request, f'Error al actualizar producto: {str(e)}')
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar producto: {str(e)}'
                })
        
        # Si llegamos aquí, hay errores
        return JsonResponse({
            'success': False,
            'errors': errores
        })
    
    # GET request - mostrar formulario
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    unidades_medida = UnidadMedida.objects.filter(activo=True).order_by('tipo', 'nombre')
    
    context = {
        'producto': producto,
        'categorias': categorias,
        'marcas': marcas,
        'unidades_medida': unidades_medida,
        'estados_producto': Producto.ESTADO_CHOICES,
    }
    return render(request, 'maestros/producto_editar.html', context)


@login_required_custom
@estado_usuario_activo
def producto_test_estado(request, pk):
    """Vista de prueba simple para cambio de estado"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        print(f"🧪 TEST: Estado recibido: '{nuevo_estado}'")
        print(f"🧪 TEST: Estado actual: '{producto.estado}'")
        
        producto.estado = nuevo_estado
        producto.save()
        
        print(f"🧪 TEST: Estado después de guardar: '{producto.estado}'")
        messages.success(request, f'Estado cambiado a {nuevo_estado}')
        return redirect('maestros:producto_test_estado', pk=pk)
    
    context = {
        'producto': producto,
        'estados_producto': Producto.ESTADO_CHOICES,
    }
    return render(request, 'maestros/producto_test_estado.html', context)


@login_required_custom
@estado_usuario_activo
@permiso_requerido('productos', 'eliminar')
def producto_eliminar(request, pk):
    """Eliminar producto con confirmación"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        try:
            print(f"[DEBUG] Intentando eliminar producto: {producto.pk} - {producto.nombre}")
            
            # Verificar si el producto tiene movimientos
            # En un sistema real verificarías inventario, compras, ventas, etc.
            nombre_producto = producto.nombre
            
            with transaction.atomic():
                producto.delete()
                print(f"[DEBUG] Producto eliminado exitosamente: {nombre_producto}")
                
            messages.success(request, f'Producto "{nombre_producto}" eliminado exitosamente.')
            response_data = {
                'success': True,
                'message': f'Producto "{nombre_producto}" eliminado exitosamente.',
                'redirect_url': reverse('maestros:producto_listar')
            }
            print(f"[DEBUG] Respuesta JSON: {response_data}")
            return JsonResponse(response_data)
        except Exception as e:
            print(f"[DEBUG] Error al eliminar producto: {str(e)}")
            messages.error(request, f'Error al eliminar producto: {str(e)}')
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar producto: {str(e)}'
            })
    
    context = {
        'producto': producto,
    }
    return render(request, 'maestros/producto_eliminar.html', context)


@csrf_exempt
def test_producto_eliminar(request, pk):
    """Test de eliminación simple sin AJAX y sin CSRF"""
    
    print(f"[DEBUG] 🚀 VISTA EJECUTADA - REQUEST: {request.method} para producto {pk}")
    print(f"[DEBUG] User: {request.user}")
    print(f"[DEBUG] POST data: {request.POST}")
    print(f"[DEBUG] Headers: {dict(request.headers)}")
    
    if request.method == 'POST':
        print(f"[DEBUG] ✅ POST REQUEST RECIBIDO")
        try:
            producto = Producto.objects.get(pk=pk)
            nombre = producto.nombre
            print(f"[DEBUG] ELIMINANDO: {pk} - {nombre}")
            
            # Eliminación directa sin transacción
            producto.delete()
            print(f"[DEBUG] ✅ ELIMINADO EXITOSAMENTE")
            
            return HttpResponse(f"✅ Producto '{nombre}' eliminado exitosamente. <a href='/maestros/productos/'>Ver lista</a>")
            
        except Producto.DoesNotExist:
            print(f"[DEBUG] ❌ Producto {pk} no encontrado")
            return HttpResponse(f"❌ Producto {pk} no encontrado")
        except Exception as e:
            print(f"[DEBUG] ❌ ERROR: {str(e)}")
            return HttpResponse(f"❌ Error: {str(e)}")
    
    # GET request
    print(f"[DEBUG] ℹ️ GET REQUEST - Mostrando formulario")
    try:
        producto = Producto.objects.get(pk=pk)
        html = f'''
        <h2>🧪 TEST Eliminar Producto (SIN CSRF)</h2>
        <p><strong>ID:</strong> {producto.pk}</p>
        <p><strong>Nombre:</strong> {producto.nombre}</p>
        <p><strong>SKU:</strong> {producto.sku}</p>
        <p><strong>Método:</strong> {request.method}</p>
        <p><strong>Usuario:</strong> {request.user}</p>
        <p style="color:orange;"><strong>⚠️ CSRF Deshabilitado para test</strong></p>
        <form method="POST">
            <button type="submit" onclick="return confirm('¿Eliminar definitivamente?')" style="background:red;color:white;padding:10px;">ELIMINAR AHORA</button>
        </form>
        <p><a href="/maestros/productos/">Volver a lista</a></p>
        '''
        return HttpResponse(html)
    except Producto.DoesNotExist:
        return HttpResponse(f"Producto {pk} no encontrado")


# ==================== PROVEEDORES ====================

@login_required_custom
@estado_usuario_activo
def proveedor_listar(request):
    """Lista de proveedores"""
    query = request.GET.get('query', '')
    estado = request.GET.get('estado', '')
    
    proveedores = Proveedor.objects.all().order_by('razon_social')
    
    if query:
        proveedores = proveedores.filter(
            Q(rut_nif__icontains=query) |
            Q(razon_social__icontains=query) |
            Q(nombre_fantasia__icontains=query)
        )
    
    if estado:
        proveedores = proveedores.filter(estado=estado)
    
    paginator = Paginator(proveedores, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'estado': estado,
    }
    return render(request, 'maestros/proveedor_listar.html', context)


@login_required_custom
@estado_usuario_activo
def proveedor_crear(request):
    """Crear nuevo proveedor"""
    if request.method == 'POST':
        errores = {}
        datos = {}
        
        try:
            # Capturar todos los datos del formulario
            datos = {
                'rut_nif': request.POST.get('rut_nif', '').strip(),
                'razon_social': request.POST.get('razon_social', '').strip(),
                'nombre_fantasia': request.POST.get('nombre_fantasia', '').strip(),
                'email': request.POST.get('email', '').strip(),
                'telefono': request.POST.get('telefono', '').strip(),
                'sitio_web': request.POST.get('sitio_web', '').strip(),
                'direccion': request.POST.get('direccion', '').strip(),
                'ciudad': request.POST.get('ciudad', '').strip(),
                'pais': request.POST.get('pais', 'Chile').strip(),
                'condiciones_pago': request.POST.get('condiciones_pago', '').strip(),
                'condiciones_pago_detalle': request.POST.get('condiciones_pago_detalle', '').strip(),
                'moneda': request.POST.get('moneda', 'CLP').strip(),
                'contacto_principal_nombre': request.POST.get('contacto_principal_nombre', '').strip(),
                'contacto_principal_email': request.POST.get('contacto_principal_email', '').strip(),
                'contacto_principal_telefono': request.POST.get('contacto_principal_telefono', '').strip(),
                'estado': request.POST.get('estado', 'ACTIVO').strip(),
                'observaciones': request.POST.get('observaciones', '').strip(),
            }
            
            # Validaciones
            if not datos['rut_nif']:
                errores['rut_nif'] = 'El RUT/NIF es obligatorio'
            else:
                # Validar formato de RUT si parece ser chileno (contiene guión)
                if '-' in datos['rut_nif']:
                    from .utils import validar_rut
                    es_valido, rut_formateado, mensaje_error = validar_rut(datos['rut_nif'])
                    if not es_valido:
                        errores['rut_nif'] = mensaje_error
                    else:
                        datos['rut_nif'] = rut_formateado  # Usar RUT formateado
                
                # Verificar unicidad
                if 'rut_nif' not in errores and Proveedor.objects.filter(rut_nif=datos['rut_nif']).exists():
                    errores['rut_nif'] = 'Ya existe un proveedor con este RUT/NIF'
                
            if not datos['razon_social']:
                errores['razon_social'] = 'La razón social es obligatoria'
                
            if not datos['email']:
                errores['email'] = 'El email es obligatorio'
            elif not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', datos['email']):
                errores['email'] = 'Email inválido'
            elif Proveedor.objects.filter(email=datos['email']).exists():
                errores['email'] = 'Ya existe un proveedor con este email'
                
            if not datos['condiciones_pago']:
                errores['condiciones_pago'] = 'Las condiciones de pago son obligatorias'
                
            # Validar sitio web si se proporciona
            if datos['sitio_web'] and not datos['sitio_web'].startswith(('http://', 'https://')):
                datos['sitio_web'] = 'https://' + datos['sitio_web']
                
            # Validar email de contacto principal si se proporciona
            if datos['contacto_principal_email'] and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', datos['contacto_principal_email']):
                errores['contacto_principal_email'] = 'Email del contacto principal inválido'
            
            if errores:
                # Si es una petición AJAX, devolver JSON con errores
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Por favor corrija los errores en el formulario',
                        'errores': errores
                    })
                
                messages.error(request, 'Por favor corrija los errores en el formulario')
                context = {
                    'errores': errores,
                    'datos': datos,
                    'CONDICIONES_PAGO_CHOICES': Proveedor.CONDICIONES_PAGO_CHOICES,
                    'ESTADO_CHOICES': Proveedor.ESTADO_CHOICES,
                }
                return render(request, 'maestros/proveedor_crear.html', context)
            
            # Crear el proveedor
            proveedor_data = {
                'rut_nif': datos['rut_nif'],
                'razon_social': datos['razon_social'],
                'email': datos['email'],
                'pais': datos['pais'],
                'condiciones_pago': datos['condiciones_pago'],
                'estado': datos['estado'],
            }
            
            # Agregar campos opcionales solo si tienen valor
            campos_opcionales = [
                'nombre_fantasia', 'telefono', 'sitio_web', 'direccion', 'ciudad',
                'condiciones_pago_detalle', 'moneda', 'contacto_principal_nombre',
                'contacto_principal_email', 'contacto_principal_telefono', 'observaciones'
            ]
            
            for campo in campos_opcionales:
                if datos[campo]:
                    proveedor_data[campo] = datos[campo]
            
            proveedor = Proveedor.objects.create(**proveedor_data)
            
            # Si es una petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Proveedor "{proveedor.razon_social}" creado exitosamente',
                    'redirect_url': reverse('maestros:proveedor_detalle', args=[proveedor.pk])
                })
            
            messages.success(request, f'Proveedor "{proveedor.razon_social}" creado exitosamente')
            return redirect('maestros:proveedor_detalle', pk=proveedor.pk)
            
        except Exception as e:
            messages.error(request, f'Error al crear proveedor: {str(e)}')
            context = {
                'errores': errores,
                'datos': datos,
                'CONDICIONES_PAGO_CHOICES': Proveedor.CONDICIONES_PAGO_CHOICES,
                'ESTADO_CHOICES': Proveedor.ESTADO_CHOICES,
            }
            return render(request, 'maestros/proveedor_crear.html', context)
    
    # GET request
    context = {
        'CONDICIONES_PAGO_CHOICES': Proveedor.CONDICIONES_PAGO_CHOICES,
        'ESTADO_CHOICES': Proveedor.ESTADO_CHOICES,
    }
    return render(request, 'maestros/proveedor_crear.html', context)


@login_required_custom
@estado_usuario_activo
def proveedor_detalle(request, pk):
    """Detalle de proveedor"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    productos = proveedor.productos.select_related('producto').filter(activo=True)[:20]
    
    context = {
        'proveedor': proveedor,
        'productos': productos,
    }
    return render(request, 'maestros/proveedor_detalle.html', context)


# ==================== CATEGORÍAS ====================

@login_required_custom
@estado_usuario_activo
def categoria_listar(request):
    """Lista de categorías"""
    categorias = Categoria.objects.all().order_by('nombre')
    
    context = {'categorias': categorias}
    return render(request, 'maestros/categoria_listar.html', context)


# ==================== MARCAS ====================

@login_required_custom
@estado_usuario_activo
def marca_listar(request):
    """Lista de marcas"""
    marcas = Marca.objects.all().order_by('nombre')
    
    context = {'marcas': marcas}
    return render(request, 'maestros/marca_listar.html', context)


# ==================== EXPORTACIÓN A EXCEL ====================

@login_required_custom
@estado_usuario_activo
def productos_exportar_excel(request):
    """Exportar productos a Excel con formato profesional"""
    
    # Aplicar los mismos filtros que en la vista de listado
    query = request.GET.get('query', '').strip()
    categoria_id = request.GET.get('categoria', '')
    marca_id = request.GET.get('marca', '')
    estado = request.GET.get('estado', '')
    orden = request.GET.get('orden', 'nombre')
    
    # Query base con relaciones
    productos = Producto.objects.select_related('categoria', 'marca', 'uom_compra', 'uom_venta', 'uom_stock').all()
    
    # Aplicar filtros
    if query:
        productos = productos.filter(
            Q(sku__icontains=query) |
            Q(nombre__icontains=query) |
            Q(categoria__nombre__icontains=query) |
            Q(marca__nombre__icontains=query)
        )
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    
    if estado:
        productos = productos.filter(estado=estado)
    
    # Aplicar ordenamiento
    orden_mapping = {
        'nombre': 'nombre',
        'nombre_desc': '-nombre',
        'sku': 'sku',
        'sku_desc': '-sku',
        'categoria': 'categoria__nombre',
        'categoria_desc': '-categoria__nombre',
        'marca': 'marca__nombre',
        'marca_desc': '-marca__nombre',
        'precio': 'precio_venta',
        'precio_desc': '-precio_venta',
        'stock': 'stock_minimo',
        'stock_desc': '-stock_minimo',
        'fecha': 'created_at',
        'fecha_desc': '-created_at',
    }
    
    if orden in orden_mapping:
        productos = productos.order_by(orden_mapping[orden])
    else:
        productos = productos.order_by('nombre')
    
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Configurar estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'SKU',
        'Nombre',
        'Categoría',
        'Marca',
        'Precio Venta',
        'Stock Mínimo',
        'UOM Compra',
        'UOM Venta',
        'UOM Stock',
        'Estado',
        'Perecible',
        'Control Lote',
        'Control Serie',
        'Fecha Creación'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        data = [
            producto.sku,
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.marca.nombre if producto.marca else '',
            float(producto.precio_venta) if producto.precio_venta else 0,
            float(producto.stock_minimo) if producto.stock_minimo else 0,
            producto.uom_compra.codigo if producto.uom_compra else '',
            producto.uom_venta.codigo if producto.uom_venta else '',
            producto.uom_stock.codigo if producto.uom_stock else '',
            producto.estado,
            'Sí' if producto.perishable else 'No',
            'Sí' if producto.control_por_lote else 'No',
            'Sí' if producto.control_por_serie else 'No',
            producto.created_at.strftime('%d/%m/%Y %H:%M') if producto.created_at else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.border = border
            
            # Aplicar alineación específica según el tipo de dato
            if col in [5, 6]:  # Precios y stock
                cell.alignment = number_alignment
                if col == 5 and value:  # Precio
                    cell.number_format = '"$"#,##0.00'
            elif col in [11, 12, 13]:  # Booleanos
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
    
    # Ajustar ancho de columnas
    column_widths = {
        1: 15,  # SKU
        2: 35,  # Nombre
        3: 20,  # Categoría
        4: 20,  # Marca
        5: 15,  # Precio
        6: 12,  # Stock
        7: 12,  # UOM Compra
        8: 12,  # UOM Venta
        9: 12,  # UOM Stock
        10: 12, # Estado
        11: 10, # Perecible
        12: 12, # Control Lote
        13: 13, # Control Serie
        14: 18  # Fecha
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Agregar información adicional
    total_productos = productos.count()
    ws.cell(row=total_productos + 3, column=1, value=f"Total de productos: {total_productos}")
    ws.cell(row=total_productos + 4, column=1, value=f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    ws.cell(row=total_productos + 5, column=1, value=f"Exportado por: {request.user.get_full_name() or request.user.username}")
    
    # Configurar respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Nombre del archivo con timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'productos_dulceria_lilis_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar workbook en la respuesta
    wb.save(response)
    
    return response


@login_required_custom
@estado_usuario_activo
def proveedor_editar(request, pk):
    """Editar proveedor existente"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        errores = {}
        datos = {}
        
        try:
            # Capturar todos los datos del formulario
            datos = {
                'rut_nif': request.POST.get('rut_nif', '').strip(),
                'razon_social': request.POST.get('razon_social', '').strip(),
                'nombre_fantasia': request.POST.get('nombre_fantasia', '').strip(),
                'email': request.POST.get('email', '').strip(),
                'telefono': request.POST.get('telefono', '').strip(),
                'sitio_web': request.POST.get('sitio_web', '').strip(),
                'direccion': request.POST.get('direccion', '').strip(),
                'ciudad': request.POST.get('ciudad', '').strip(),
                'pais': request.POST.get('pais', 'Chile').strip(),
                'condiciones_pago': request.POST.get('condiciones_pago', '').strip(),
                'condiciones_pago_detalle': request.POST.get('condiciones_pago_detalle', '').strip(),
                'moneda': request.POST.get('moneda', 'CLP').strip(),
                'contacto_principal_nombre': request.POST.get('contacto_principal_nombre', '').strip(),
                'contacto_principal_email': request.POST.get('contacto_principal_email', '').strip(),
                'contacto_principal_telefono': request.POST.get('contacto_principal_telefono', '').strip(),
                'estado': request.POST.get('estado', 'ACTIVO').strip(),
                'observaciones': request.POST.get('observaciones', '').strip(),
            }
            
            # Validaciones
            if not datos['rut_nif']:
                errores['rut_nif'] = 'El RUT/NIF es obligatorio'
            else:
                # Validar formato de RUT si parece ser chileno (contiene guión)
                if '-' in datos['rut_nif']:
                    from .utils import validar_rut
                    es_valido, rut_formateado, mensaje_error = validar_rut(datos['rut_nif'])
                    if not es_valido:
                        errores['rut_nif'] = mensaje_error
                    else:
                        datos['rut_nif'] = rut_formateado  # Usar RUT formateado
                
                # Verificar unicidad
                if 'rut_nif' not in errores and Proveedor.objects.filter(rut_nif=datos['rut_nif']).exclude(pk=proveedor.pk).exists():
                    errores['rut_nif'] = 'Ya existe un proveedor con este RUT/NIF'
                
            if not datos['razon_social']:
                errores['razon_social'] = 'La razón social es obligatoria'
                
            if not datos['email']:
                errores['email'] = 'El email es obligatorio'
            elif '@' not in datos['email']:
                errores['email'] = 'Ingrese un email válido'
            elif Proveedor.objects.filter(email=datos['email']).exclude(pk=proveedor.pk).exists():
                errores['email'] = 'Ya existe un proveedor con este email'
                
            if not datos['condiciones_pago']:
                errores['condiciones_pago'] = 'Las condiciones de pago son obligatorias'
            
            # Si hay errores, devolver respuesta con errores
            if errores:
                return JsonResponse({
                    'success': False,
                    'errors': errores,
                    'message': 'Por favor corrija los errores indicados'
                })
            
            # Actualizar proveedor
            for field, value in datos.items():
                if value or field in ['observaciones', 'nombre_fantasia', 'telefono', 'sitio_web', 'direccion', 'ciudad', 'condiciones_pago_detalle', 'contacto_principal_nombre', 'contacto_principal_email', 'contacto_principal_telefono']:
                    setattr(proveedor, field, value)
            
            proveedor.save()
            
            messages.success(request, f'Proveedor "{proveedor.razon_social}" actualizado exitosamente.')
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{proveedor.razon_social}" actualizado exitosamente.',
                'redirect_url': reverse('maestros:proveedor_detalle', kwargs={'pk': proveedor.pk})
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar proveedor: {str(e)}'
            })
    
    # GET request - mostrar formulario de edición
    context = {
        'proveedor': proveedor,
        'condiciones_pago_choices': Proveedor.CONDICIONES_PAGO_CHOICES,
        'estado_choices': Proveedor.ESTADO_CHOICES,
    }
    return render(request, 'maestros/proveedor_editar.html', context)


@login_required_custom
@estado_usuario_activo
def proveedor_eliminar(request, pk):
    """Eliminar proveedor con confirmación"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        try:
            # Verificar si el proveedor tiene productos asociados
            productos_asociados = proveedor.productos.count()
            
            if productos_asociados > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'No se puede eliminar el proveedor porque tiene {productos_asociados} productos asociados. Primero debe eliminar o reasignar estos productos.'
                })
            
            nombre_proveedor = proveedor.razon_social
            proveedor.delete()
            
            messages.success(request, f'Proveedor "{nombre_proveedor}" eliminado exitosamente.')
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{nombre_proveedor}" eliminado exitosamente.',
                'redirect_url': reverse('maestros:proveedor_listar')
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar proveedor: {str(e)}'
            })
    
    # GET request - mostrar confirmación
    # Verificar productos asociados para mostrar advertencia
    productos_asociados = proveedor.productos.count()
    
    context = {
        'proveedor': proveedor,
        'productos_asociados': productos_asociados,
    }
    return render(request, 'maestros/proveedor_eliminar.html', context)

