from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from autenticacion.decorators import permission_required, role_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import json

from maestros.models import Categoria, Marca, UnidadMedida, Producto
from inventario.models import Bodega, StockActual


@login_required
@permission_required('productos.ver')
def lista_productos(request):
    """Vista para listar productos con búsqueda, paginación y ordenamiento"""
    
    # Obtener parámetros de búsqueda y filtros
    busqueda = request.GET.get('busqueda', '')
    categoria_id = request.GET.get('categoria', '')
    marca_id = request.GET.get('marca', '')
    estado = request.GET.get('estado', '')
    ordenar_por = request.GET.get('ordenar', 'nombre')
    direccion = request.GET.get('direccion', 'asc')
    
    # Obtener paginación de la sesión o usar valor por defecto
    items_por_pagina = request.GET.get('items_por_pagina', request.session.get('items_por_pagina', 15))
    request.session['items_por_pagina'] = int(items_por_pagina)
    
    # Consulta base
    productos = Producto.objects.select_related('categoria', 'marca').all()
    
    # Aplicar filtros de búsqueda
    if busqueda:
        productos = productos.filter(
            Q(sku__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    
    if estado:
        productos = productos.filter(estado=estado)
    
    # Aplicar ordenamiento
    campos_ordenamiento = {
        'sku': 'sku',
        'nombre': 'nombre',
        'categoria': 'categoria__nombre',
        'marca': 'marca__nombre',
        'precio_venta': 'precio_venta',
        'costo_estandar': 'costo_estandar',
        'estado': 'estado',
        'created_at': 'created_at'
    }
    
    campo_orden = campos_ordenamiento.get(ordenar_por, 'nombre')
    if direccion == 'desc':
        campo_orden = f'-{campo_orden}'
    
    productos = productos.order_by(campo_orden)
    
    # Paginación
    paginator = Paginator(productos, int(items_por_pagina))
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Datos para los filtros
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    
    # Estadísticas generales
    total_productos = Producto.objects.count()
    productos_activos = Producto.objects.filter(estado='ACTIVO').count()
    total_categorias = Categoria.objects.filter(activo=True).count()
    total_marcas = Marca.objects.filter(activo=True).count()
    
    # Verificar permisos del usuario
    user_permissions = {
        'puede_crear': _check_permission(request.user, 'productos.crear'),
        'puede_editar': _check_permission(request.user, 'productos.editar'),
        'puede_eliminar': _check_permission(request.user, 'productos.eliminar'),
        'puede_exportar': _check_permission(request.user, 'productos.exportar'),
    }
    
    context = {
        'page_obj': page_obj,
        'busqueda': busqueda,
        'categoria_id': int(categoria_id) if categoria_id else '',
        'marca_id': int(marca_id) if marca_id else '',
        'estado': estado,
        'ordenar_por': ordenar_por,
        'direccion': direccion,
        'items_por_pagina': int(items_por_pagina),
        'categorias': categorias,
        'marcas': marcas,
        'estados': Producto.ESTADO_CHOICES,
        'total_productos': paginator.count,
        'productos_filtrados': paginator.count,
        'user_permissions': user_permissions,
        # Estadísticas
        'stats': {
            'total_productos': total_productos,
            'productos_activos': productos_activos,
            'total_categorias': total_categorias,
            'total_marcas': total_marcas,
        }
    }
    
    return render(request, 'productos/lista.html', context)


@login_required
@permission_required('productos.crear')
def crear_producto(request):
    """Vista para crear un nuevo producto"""
    
    if request.method == 'POST':
        try:
            # Debug: imprimir datos recibidos
            print("=== DEBUG CREAR PRODUCTO ===")
            print("POST data:", dict(request.POST))
            print("Bodega inicial:", request.POST.get('bodega_inicial'))
            print("Cantidad inicial:", request.POST.get('cantidad_inicial'))
            print("===========================")
            
            # Validar datos obligatorios
            sku = request.POST.get('sku', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            categoria_id = request.POST.get('categoria')
            precio_venta = request.POST.get('precio_venta')
            costo_estandar = request.POST.get('costo_estandar', '') or None
            
            if not all([sku, nombre, categoria_id]):
                return JsonResponse({
                    'success': False,
                    'message': 'Los campos SKU, nombre y categoría son obligatorios'
                })
            
            # Validar precio de venta
            try:
                precio_venta = float(precio_venta)
                if precio_venta <= 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'El precio de venta debe ser mayor que 0'
                    })
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'El precio de venta debe ser un número válido'
                })
            
            # Validar costo estándar si se proporciona
            if costo_estandar:
                try:
                    costo_estandar = float(costo_estandar)
                    if costo_estandar <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'El costo estándar debe ser mayor que 0'
                        })
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False,
                        'message': 'El costo estándar debe ser un número válido'
                    })
            
            # Verificar que el SKU no exista
            if Producto.objects.filter(sku=sku).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Ya existe un producto con el SKU: {sku}'
                })
            
            # Obtener objetos relacionados
            categoria = get_object_or_404(Categoria, id=categoria_id)
            marca_id = request.POST.get('marca') or None
            marca = get_object_or_404(Marca, id=marca_id) if marca_id else None
            
            # Crear el producto
            producto = Producto.objects.create(
                sku=sku,
                nombre=nombre,
                descripcion=request.POST.get('descripcion', '').strip(),
                categoria=categoria,
                marca=marca,
                modelo=request.POST.get('modelo', '').strip(),
                precio_venta=precio_venta,
                costo_estandar=costo_estandar,
                impuesto_iva=float(request.POST.get('impuesto_iva', 19)),
                estado=request.POST.get('estado', 'ACTIVO'),
                perishable=request.POST.get('perishable') == 'on',
                control_por_lote=request.POST.get('control_por_lote') == 'on',
                control_por_serie=request.POST.get('control_por_serie') == 'on',
                imagen_url=request.POST.get('imagen_url', '').strip() or None,
            )
            
            # Asignar stock inicial a bodega específica si se proporciona
            bodega_id = request.POST.get('bodega_inicial')
            cantidad_inicial = request.POST.get('cantidad_inicial', '0')
            
            if bodega_id:
                try:
                    bodega = Bodega.objects.get(id=bodega_id, activo=True)
                    cantidad = float(cantidad_inicial) if cantidad_inicial else 0
                    
                    # Esperar a que las signals creen el stock (o crearlo si no existe)
                    from django.db import transaction
                    import time
                    time.sleep(0.1)  # Pequeña pausa para que la signal se ejecute
                    
                    # Intentar obtener o crear el stock en la bodega seleccionada
                    stock, created = StockActual.objects.get_or_create(
                        producto=producto,
                        bodega=bodega,
                        defaults={
                            'cantidad_disponible': cantidad,
                            'cantidad_reservada': 0,
                            'cantidad_transito': 0
                        }
                    )
                    
                    if not created and cantidad > 0:
                        # Si ya existía, actualizar la cantidad
                        stock.cantidad_disponible = cantidad
                        stock.save()
                    
                    mensaje = f'Producto "{producto.nombre}" creado exitosamente'
                    if cantidad > 0:
                        mensaje += f' con {cantidad} unidades en {bodega.nombre}'
                    
                except Bodega.DoesNotExist:
                    mensaje = f'Producto "{producto.nombre}" creado, pero bodega no encontrada'
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    mensaje = f'Producto "{producto.nombre}" creado, pero error al asignar stock: {str(e)}'
            else:
                mensaje = f'Producto "{producto.nombre}" creado exitosamente'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'producto_id': producto.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear el producto: {str(e)}'
            })
    
    # GET request - mostrar formulario
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    unidades = UnidadMedida.objects.filter(activo=True).order_by('nombre')
    bodegas = Bodega.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'categorias': categorias,
        'marcas': marcas,
        'unidades': unidades,
        'bodegas': bodegas,
        'estados': Producto.ESTADO_CHOICES,
    }
    
    return render(request, 'productos/crear.html', context)


@login_required
@permission_required('productos.editar')
def editar_producto(request, producto_id):
    """Vista para editar un producto existente"""
    
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        try:
            # Validar datos obligatorios
            sku = request.POST.get('sku', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            categoria_id = request.POST.get('categoria')
            precio_venta = request.POST.get('precio_venta')
            costo_estandar = request.POST.get('costo_estandar', '') or None
            
            if not all([sku, nombre, categoria_id]):
                return JsonResponse({
                    'success': False,
                    'message': 'Los campos SKU, nombre y categoría son obligatorios'
                })
            
            # Validar precio de venta
            try:
                precio_venta = float(precio_venta)
                if precio_venta <= 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'El precio de venta debe ser mayor que 0'
                    })
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'message': 'El precio de venta debe ser un número válido'
                })
            
            # Validar costo estándar si se proporciona
            if costo_estandar:
                try:
                    costo_estandar = float(costo_estandar)
                    if costo_estandar <= 0:
                        return JsonResponse({
                            'success': False,
                            'message': 'El costo estándar debe ser mayor que 0'
                        })
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False,
                        'message': 'El costo estándar debe ser un número válido'
                    })
            
            # Verificar que el SKU no exista en otro producto
            if Producto.objects.filter(sku=sku).exclude(id=producto.id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Ya existe otro producto con el SKU: {sku}'
                })
            
            # Obtener objetos relacionados
            categoria = get_object_or_404(Categoria, id=categoria_id)
            marca_id = request.POST.get('marca') or None
            marca = get_object_or_404(Marca, id=marca_id) if marca_id else None
            
            # Actualizar el producto
            producto.sku = sku
            producto.nombre = nombre
            producto.descripcion = request.POST.get('descripcion', '').strip()
            producto.categoria = categoria
            producto.marca = marca
            producto.modelo = request.POST.get('modelo', '').strip()
            producto.precio_venta = precio_venta
            producto.costo_estandar = costo_estandar
            producto.impuesto_iva = float(request.POST.get('impuesto_iva', 19))
            producto.estado = request.POST.get('estado', 'ACTIVO')
            producto.perishable = request.POST.get('perishable') == 'on'
            producto.control_por_lote = request.POST.get('control_por_lote') == 'on'
            producto.control_por_serie = request.POST.get('control_por_serie') == 'on'
            producto.imagen_url = request.POST.get('imagen_url', '').strip() or None
            
            producto.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Producto "{producto.nombre}" actualizado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar el producto: {str(e)}'
            })
    
    # GET request - mostrar formulario
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    unidades = UnidadMedida.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'producto': producto,
        'categorias': categorias,
        'marcas': marcas,
        'unidades': unidades,
        'estados': Producto.ESTADO_CHOICES,
    }
    
    return render(request, 'productos/editar.html', context)


@login_required
@permission_required('productos.ver')
def ver_producto(request, producto_id):
    """Vista para ver detalles de un producto"""
    
    producto = get_object_or_404(Producto, id=producto_id)
    
    # Verificar permisos del usuario
    user_permissions = {
        'puede_editar': _check_permission(request.user, 'productos.editar'),
        'puede_eliminar': _check_permission(request.user, 'productos.eliminar'),
    }
    
    context = {
        'producto': producto,
        'user_permissions': user_permissions,
    }
    
    return render(request, 'productos/detalle.html', context)


@login_required
@permission_required('productos.eliminar')
@require_http_methods(["POST"])
def eliminar_producto(request, producto_id):
    """Vista para eliminar un producto"""
    
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        nombre_producto = producto.nombre
        
        # Verificar si el producto tiene relaciones que impidan eliminarlo
        # Aquí puedes agregar validaciones adicionales según tus reglas de negocio
        
        producto.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Producto "{nombre_producto}" eliminado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar el producto: {str(e)}'
        })


@login_required
@permission_required('productos.exportar')
def exportar_productos(request):
    """Vista para exportar productos a Excel"""
    
    try:
        # Obtener los mismos filtros que la lista
        busqueda = request.GET.get('busqueda', '')
        categoria_id = request.GET.get('categoria', '')
        marca_id = request.GET.get('marca', '')
        estado = request.GET.get('estado', '')
        ordenar_por = request.GET.get('ordenar', 'nombre')
        direccion = request.GET.get('direccion', 'asc')
        
        # Aplicar filtros
        productos = Producto.objects.select_related('categoria', 'marca').all()
        
        if busqueda:
            productos = productos.filter(
                Q(sku__icontains=busqueda) |
                Q(nombre__icontains=busqueda) |
                Q(descripcion__icontains=busqueda)
            )
        
        if categoria_id:
            productos = productos.filter(categoria_id=categoria_id)
        
        if marca_id:
            productos = productos.filter(marca_id=marca_id)
        
        if estado:
            productos = productos.filter(estado=estado)
        
        # Aplicar ordenamiento
        campos_ordenamiento = {
            'sku': 'sku',
            'nombre': 'nombre',
            'categoria': 'categoria__nombre',
            'marca': 'marca__nombre',
            'precio_venta': 'precio_venta',
            'costo_estandar': 'costo_estandar',
            'estado': 'estado',
            'created_at': 'created_at'
        }
        
        campo_orden = campos_ordenamiento.get(ordenar_por, 'nombre')
        if direccion == 'desc':
            campo_orden = f'-{campo_orden}'
        
        productos = productos.order_by(campo_orden)
        
        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'SKU', 'Nombre', 'Descripción', 'Categoría', 'Marca', 'Modelo',
            'Precio Venta', 'Costo Estándar', 'IVA %', 'Estado', 
            'Perecedero', 'Control Lote', 'Control Serie', 'Fecha Creación'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row, producto in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=producto.sku)
            ws.cell(row=row, column=2, value=producto.nombre)
            ws.cell(row=row, column=3, value=producto.descripcion or '')
            ws.cell(row=row, column=4, value=producto.categoria.nombre if producto.categoria else '')
            ws.cell(row=row, column=5, value=producto.marca.nombre if producto.marca else '')
            ws.cell(row=row, column=6, value=producto.modelo or '')
            ws.cell(row=row, column=7, value=float(producto.precio_venta) if producto.precio_venta else 0)
            ws.cell(row=row, column=8, value=float(producto.costo_estandar) if producto.costo_estandar else 0)
            ws.cell(row=row, column=9, value=float(producto.impuesto_iva))
            ws.cell(row=row, column=10, value=producto.estado)
            ws.cell(row=row, column=11, value='Sí' if producto.perishable else 'No')
            ws.cell(row=row, column=12, value='Sí' if producto.control_por_lote else 'No')
            ws.cell(row=row, column=13, value='Sí' if producto.control_por_serie else 'No')
            ws.cell(row=row, column=14, value=producto.created_at.strftime('%d/%m/%Y %H:%M'))
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'productos_dulceria_lilis_{fecha_actual}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar productos: {str(e)}')
        return redirect('productos:lista')


def _check_permission(user, permission_name):
    """Función auxiliar para verificar permisos"""
    if not user.is_authenticated:
        return False
    
    user_role = getattr(user, 'rol', None)
    if not user_role:
        return False
    
    # Los administradores tienen acceso completo
    if user_role.nombre == 'Administrador':
        return True
    
    if not user_role.permisos:
        return False
    
    # Separar el permiso en módulo y acción
    parts = permission_name.split('.')
    if len(parts) != 2:
        return False
    
    modulo, accion = parts
    
    # Verificar si el módulo existe y tiene el permiso
    if modulo in user_role.permisos and accion in user_role.permisos[modulo]:
        return user_role.permisos[modulo][accion]
    
    return False
