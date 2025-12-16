import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from .models import Usuario, PasswordResetToken, PasswordChangeCode
from .forms import (
    EditarPerfilForm, CambiarPasswordForm, RecuperarPasswordForm, 
    ResetearPasswordForm, SolicitarCodigoCambioForm, VerificarCodigoCambioForm,
    RegistroUsuarioForm
)
from .utils import crear_token_reset, enviar_email_reset_password, validar_token_reset, marcar_token_usado, procesar_avatar, crear_codigo_cambio_password, enviar_email_codigo_cambio, validar_codigo_cambio_password, marcar_codigo_usado


@csrf_protect
def registro_view(request):
    """Vista de registro de nuevos usuarios"""
    if request.user.is_authenticated:
        return redirect('autenticacion:dashboard')
    
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            # Guardar nuevo usuario
            user = form.save()
            
            # Mensaje de éxito
            messages.success(
                request,
                f'¡Cuenta creada exitosamente! Bienvenido, {user.get_full_name()}. '
                'Ya puedes iniciar sesión.'
            )
            
            # Respuesta AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Cuenta creada exitosamente',
                    'redirect_url': '/auth/login/'
                })
            
            return redirect('autenticacion:login')
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            # Mostrar errores en mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = RegistroUsuarioForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'autenticacion/registro.html', context)


def login_view(request):
    """Vista de inicio de sesión con cycle_key para seguridad y bloqueo por intentos fallidos"""
    if request.user.is_authenticated:
        return redirect('autenticacion:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Intentar obtener el usuario por username
        usuario = None
        try:
            usuario = Usuario.objects.get(username=username)
        except Usuario.DoesNotExist:
            # Usuario no existe - mensaje genérico sin revelar que no existe
            messages.error(request, 'Usuario o contraseña incorrectos.')
            return render(request, 'autenticacion/login.html', {'username': username})
        
        # Si llegamos aquí, el usuario existe
        # Verificar si la cuenta está bloqueada temporalmente
        if usuario.esta_bloqueado():
            minutos_restantes = usuario.tiempo_restante_bloqueo()
            messages.error(
                request, 
                f'Tu cuenta está bloqueada temporalmente por múltiples intentos fallidos. '
                f'Podrás intentar nuevamente en {minutos_restantes} minuto(s).'
            )
            return render(request, 'autenticacion/login.html', {'username': username})
        
        # Si el bloqueo expiró, resetear contadores
        if usuario.bloqueado_hasta and timezone.now() >= usuario.bloqueado_hasta:
            usuario.intentos_fallidos = 0
            usuario.bloqueado_hasta = None
            usuario.save(update_fields=['intentos_fallidos', 'bloqueado_hasta'])
        
        # Autenticar usuario
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Verificar estado del usuario
            if user.estado == 'BLOQUEADO':
                messages.error(request, 'Tu cuenta está bloqueada permanentemente. Contacta al administrador.')
                return render(request, 'autenticacion/login.html', {'username': username})
            
            if user.estado != 'ACTIVO':
                messages.error(request, 'Tu cuenta no está activa. Contacta al administrador.')
                return render(request, 'autenticacion/login.html', {'username': username})
            
            # Login exitoso - resetear intentos fallidos
            if user.intentos_fallidos > 0 or user.bloqueado_hasta:
                user.intentos_fallidos = 0
                user.bloqueado_hasta = None
                user.save(update_fields=['intentos_fallidos', 'bloqueado_hasta'])
            
            # Regenerar la clave de sesión para prevenir session fixation
            request.session.cycle_key()
            
            # Iniciar sesión
            login(request, user)
            
            # Actualizar último acceso
            user.ultimo_acceso = timezone.now()
            user.save(update_fields=['ultimo_acceso'])
            
            # Inicializar variables de sesión
            request.session['login_time'] = timezone.now().isoformat()
            request.session['carrito'] = []
            request.session['notificaciones'] = []
            
            # Mensaje de bienvenida
            messages.success(request, f'¡Bienvenido, {user.get_full_name()}!')
            
            # Agregar notificación de bienvenida
            request.session['notificaciones'] = [{
                'id': 1,
                'titulo': 'Bienvenido',
                'mensaje': f'Hola {user.get_full_name()}, has iniciado sesión exitosamente.',
                'tipo': 'success',
                'leida': False,
                'tiempo': timezone.now().isoformat()
            }]
            
            # Redirigir a la página solicitada o al dashboard
            next_url = request.GET.get('next', 'autenticacion:dashboard')
            return redirect(next_url)
        else:
            # Contraseña incorrecta (usuario existe pero password mal)
            usuario.intentos_fallidos += 1
            
            # Configuración de bloqueo
            MAX_INTENTOS = 3
            TIEMPO_BLOQUEO_MINUTOS = 15
            
            if usuario.intentos_fallidos >= MAX_INTENTOS:
                # Bloquear cuenta temporalmente
                usuario.bloqueado_hasta = timezone.now() + timezone.timedelta(minutes=TIEMPO_BLOQUEO_MINUTOS)
                usuario.save(update_fields=['intentos_fallidos', 'bloqueado_hasta'])
                
                messages.error(
                    request,
                    f'Has superado el número máximo de intentos ({MAX_INTENTOS}). '
                    f'Tu cuenta ha sido bloqueada temporalmente por {TIEMPO_BLOQUEO_MINUTOS} minutos.'
                )
            else:
                # Informar intentos restantes
                intentos_restantes = MAX_INTENTOS - usuario.intentos_fallidos
                usuario.save(update_fields=['intentos_fallidos'])
                
                messages.error(
                    request,
                    f'Contraseña incorrecta. Te quedan {intentos_restantes} intento(s) antes de que tu cuenta sea bloqueada temporalmente.'
                )
            
            # Preservar el username en el formulario
            return render(request, 'autenticacion/login.html', {'username': username})
    
    return render(request, 'autenticacion/login.html')


def logout_view(request):
    """Vista para cerrar sesión"""
    if request.user.is_authenticated:
        username = request.user.get_full_name()
        logout(request)
        messages.info(request, f'Hasta pronto, {username}.')
    
    return redirect('autenticacion:login')


def dashboard(request):
    """Vista del dashboard principal"""
    if not request.user.is_authenticated:
        return redirect('autenticacion:login')
    
    # Importar modelos para estadísticas
    from django.db import models
    from maestros.models import Producto, Proveedor
    from inventario.models import MovimientoInventario, StockActual, AlertaStock
    from autenticacion.models import Usuario
    
    # Estadísticas generales
    stats = {
        'productos_total': Producto.objects.filter(estado='ACTIVO').count(),
        'productos_bajo_stock': Producto.objects.filter(
            stock_actual__lte=models.F('stock_minimo'),
            estado='ACTIVO'
        ).count() if hasattr(Producto, 'stock_actual') else 0,
        'proveedores_activos': Proveedor.objects.filter(estado='ACTIVO').count(),
        'usuarios_activos': Usuario.objects.filter(estado='ACTIVO').count(),
        'movimientos_hoy': MovimientoInventario.objects.filter(
            fecha_movimiento__date=timezone.now().date()
        ).count(),
        'alertas_pendientes': AlertaStock.objects.filter(estado='ACTIVA').count(),
    }
    
    # Movimientos recientes
    movimientos_recientes = MovimientoInventario.objects.select_related(
        'producto', 'usuario'
    ).order_by('-fecha_movimiento')[:5]
    
    # Alertas críticas
    alertas_criticas = AlertaStock.objects.filter(
        estado='ACTIVA',
        prioridad__in=['CRITICA', 'ALTA']
    ).select_related('producto', 'bodega').order_by('-fecha_generacion')[:5]
    
    # Productos más movidos
    productos_movidos = MovimientoInventario.objects.values(
        'producto__nombre'
    ).annotate(
        total_movimientos=models.Count('id')
    ).order_by('-total_movimientos')[:5]
    
    context = {
        'usuario': request.user,
        'stats': stats,
        'movimientos_recientes': movimientos_recientes,
        'alertas_criticas': alertas_criticas,
        'productos_movidos': productos_movidos,
        'carrito_count': len(request.session.get('carrito', [])),
        'notificaciones_count': len([n for n in request.session.get('notificaciones', []) if not n.get('leida', True)]),
    }
    
    return render(request, 'autenticacion/dashboard.html', context)


@login_required
def perfil_usuario(request):
    """Vista del perfil del usuario - disponible para todos los usuarios autenticados"""
    
    # Verificar que el usuario esté activo
    if request.user.estado != 'ACTIVO':
        messages.error(request, 'Tu cuenta no está activa. Contacta al administrador.')
        return redirect('autenticacion:login')
    
    context = {
        'usuario': request.user,
        'carrito_count': len(request.session.get('carrito', [])),
        'notificaciones_count': len([n for n in request.session.get('notificaciones', []) if not n.get('leida', True)]),
        'puede_editar_perfil': True,  # Todos los usuarios autenticados pueden editar su propio perfil
        'es_lector': request.user.rol and request.user.rol.nombre == 'Lector',
        'es_editor': request.user.rol and request.user.rol.nombre == 'Editor',
        'es_admin': request.user.rol and request.user.rol.nombre == 'Administrador',
    }
    
    return render(request, 'autenticacion/perfil.html', context)


@login_required
@csrf_protect
def editar_perfil(request):
    """Vista para editar el perfil del usuario - disponible para todos los usuarios autenticados"""
    
    # Verificar que el usuario esté activo
    if request.user.estado != 'ACTIVO':
        messages.error(request, 'Tu cuenta no está activa. Contacta al administrador.')
        return redirect('autenticacion:login')
    
    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                
                # Procesar avatar si se subió uno nuevo
                if 'avatar' in request.FILES:
                    avatar_procesado = procesar_avatar(request.FILES['avatar'], usuario)
                    if avatar_procesado:
                        usuario.avatar = avatar_procesado
                
                # Mantener campos que el usuario no debería cambiar
                usuario.username = request.user.username
                usuario.rol = request.user.rol
                usuario.estado = request.user.estado
                usuario.is_staff = request.user.is_staff
                usuario.is_superuser = request.user.is_superuser
                
                usuario.save()
                messages.success(request, '¡Tu perfil ha sido actualizado exitosamente!')
                
                # Respuesta AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Perfil actualizado exitosamente',
                        'avatar_url': usuario.avatar.url if usuario.avatar else None,
                        'redirect_url': reverse('autenticacion:perfil_usuario')
                    })
                
                return redirect('autenticacion:perfil_usuario')
                
            except Exception as e:
                error_msg = f'Error al actualizar el perfil: {str(e)}'
                messages.error(request, error_msg)
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    })
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EditarPerfilForm(instance=request.user)
    
    context = {
        'form': form,
        'usuario': request.user,
        'carrito_count': len(request.session.get('carrito', [])),
        'notificaciones_count': len([n for n in request.session.get('notificaciones', []) if not n.get('leida', True)]),
        'puede_editar_perfil': True,  # Todos los usuarios autenticados pueden editar su propio perfil
        'es_lector': request.user.rol and request.user.rol.nombre == 'Lector',
        'es_editor': request.user.rol and request.user.rol.nombre == 'Editor',
        'es_admin': request.user.rol and request.user.rol.nombre == 'Administrador',
    }
    
    return render(request, 'autenticacion/editar_perfil.html', context)


@login_required
@csrf_protect
def solicitar_codigo_cambio(request):
    """Vista para solicitar código de verificación para cambio de contraseña"""
    if request.method == 'POST':
        form = SolicitarCodigoCambioForm(request.user, request.POST)
        if form.is_valid():
            # Obtener IP del usuario
            ip_address = request.META.get('HTTP_X_FORWARDED_FOR', 
                                        request.META.get('REMOTE_ADDR'))
            if ip_address:
                ip_address = ip_address.split(',')[0].strip()
            
            # Crear código de verificación
            codigo_cambio = crear_codigo_cambio_password(request.user, ip_address)
            
            # Enviar email con código
            email_enviado = enviar_email_codigo_cambio(request.user, codigo_cambio, request)
            
            if email_enviado:
                messages.success(
                    request, 
                    f'Se ha enviado un código de verificación a {request.user.email}. '
                    'El código es válido por 10 minutos.'
                )
            else:
                # Mostrar código temporalmente cuando no se puede enviar email
                messages.warning(
                    request, 
                    f'No se pudo enviar el email. Tu código de verificación es: {codigo_cambio.codigo}. '
                    'Válido por 10 minutos. (Configurar email SMTP para envío automático)'
                )
            
            # Respuesta AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Código generado exitosamente',
                    'redirect_url': '/auth/verificar-codigo-cambio/'
                })
            
            return redirect('autenticacion:verificar_codigo_cambio')
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for error in form.errors.values():
                messages.error(request, error[0])
    else:
        form = SolicitarCodigoCambioForm(request.user)
    
    context = {
        'form': form,
        'usuario': request.user,
        'carrito_count': len(request.session.get('carrito', [])),
        'notificaciones_count': len([n for n in request.session.get('notificaciones', []) if not n.get('leida', True)]),
    }
    
    return render(request, 'autenticacion/solicitar_codigo_cambio.html', context)


@login_required
@csrf_protect
def verificar_codigo_cambio(request):
    """Vista para verificar código y cambiar contraseña"""
    if request.method == 'POST':
        form = VerificarCodigoCambioForm(request.POST)
        if form.is_valid():
            codigo = form.cleaned_data['codigo']
            nueva_password = form.cleaned_data['new_password1']
            
            # Validar código
            codigo_cambio = validar_codigo_cambio_password(codigo)
            
            if codigo_cambio and codigo_cambio.usuario == request.user:
                # Cambiar contraseña
                request.user.set_password(nueva_password)
                request.user.save()
                
                # Marcar código como usado
                marcar_codigo_usado(codigo_cambio)
                
                # Mantener sesión activa
                update_session_auth_hash(request, request.user)
                
                messages.success(
                    request, 
                    '¡Tu contraseña ha sido cambiada exitosamente!'
                )
                
                # Respuesta AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Contraseña cambiada exitosamente',
                        'redirect_url': '/auth/perfil/'
                    })
                
                return redirect('autenticacion:perfil_usuario')
            else:
                error_msg = 'El código es inválido o ha expirado.'
                messages.error(request, error_msg)
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    })
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for error in form.errors.values():
                messages.error(request, error[0])
    else:
        form = VerificarCodigoCambioForm()
    
    context = {
        'form': form,
        'usuario': request.user,
        'carrito_count': len(request.session.get('carrito', [])),
        'notificaciones_count': len([n for n in request.session.get('notificaciones', []) if not n.get('leida', True)]),
    }
    
    return render(request, 'autenticacion/verificar_codigo_cambio.html', context)


# Vista legacy para cambiar contraseña (mantenida para compatibilidad)
@login_required  
@csrf_protect
def cambiar_password(request):
    """Vista legacy - redirige al nuevo flujo"""
    return redirect('autenticacion:solicitar_codigo_cambio')


@csrf_protect
def recuperar_password(request):
    """Vista para solicitar código de verificación para recuperar contraseña (sin estar logueado)"""
    if request.user.is_authenticated:
        return redirect('autenticacion:dashboard')
    
    if request.method == 'POST':
        form = RecuperarPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                usuario = Usuario.objects.get(email=email, estado='ACTIVO')
                
                # Crear código de verificación (esto internamente invalida códigos anterior)
                codigo_cambio = crear_codigo_cambio_password(
                    usuario, 
                    request.META.get('REMOTE_ADDR', '127.0.0.1')
                )
                
                if codigo_cambio:
                    # Enviar email con código
                    email_enviado = enviar_email_codigo_cambio(usuario, codigo_cambio, request)
                    
                    if email_enviado:
                        # Guardar el usuario en sesión para la siguiente vista
                        request.session['recovery_user_id'] = usuario.id
                        
                        messages.success(
                            request, 
                            f'Se ha enviado un código de verificación a {email}. '
                            'Revisa tu bandeja de entrada y spam.'
                        )
                    else:
                        # Mostrar código temporalmente cuando no se puede enviar email
                        request.session['recovery_user_id'] = usuario.id
                        
                        messages.warning(
                            request, 
                            f'No se pudo enviar el email. El código de verificación es: {codigo_cambio.codigo}. '
                            'Válido por 10 minutos. (Configurar email SMTP para envío automático)'
                        )
                    
                    # Respuesta AJAX
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'message': f'Código generado para {email}',
                            'redirect_url': '/auth/verificar-codigo-recuperacion/'
                        })
                    
                    return redirect('autenticacion:verificar_codigo_recuperacion')
                else:
                    messages.error(request, 'Error generando el código de verificación.')
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': 'Error generando código'
                        })
                
            except Usuario.DoesNotExist:
                messages.error(request, 'No existe una cuenta activa asociada a este email.')
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Email no encontrado'
                    })
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for error in form.errors.values():
                messages.error(request, error[0])
    else:
        form = RecuperarPasswordForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'autenticacion/recuperar_password.html', context)


@csrf_protect
def verificar_codigo_recuperacion(request):
    """Vista para verificar código y cambiar contraseña (sin estar logueado)"""
    if request.user.is_authenticated:
        return redirect('autenticacion:dashboard')
    
    # Verificar que hay un usuario en sesión para recuperación
    user_id = request.session.get('recovery_user_id')
    if not user_id:
        messages.error(request, 'Sesión de recuperación expirada. Solicita un nuevo código.')
        return redirect('autenticacion:recuperar_password')
    
    try:
        usuario = Usuario.objects.get(id=user_id, estado='ACTIVO')
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado.')
        return redirect('autenticacion:recuperar_password')
    
    if request.method == 'POST':
        form = VerificarCodigoCambioForm(request.POST)
        if form.is_valid():
            codigo = form.cleaned_data['codigo']
            nueva_password = form.cleaned_data['new_password1']
            
            # Validar código
            codigo_cambio = validar_codigo_cambio_password(codigo)
            
            if codigo_cambio and codigo_cambio.usuario == usuario:
                # Cambiar contraseña
                usuario.set_password(nueva_password)
                usuario.save()
                
                # Marcar código como usado
                marcar_codigo_usado(codigo_cambio)
                
                # Limpiar sesión de recuperación
                if 'recovery_user_id' in request.session:
                    del request.session['recovery_user_id']
                
                messages.success(
                    request, 
                    '¡Tu contraseña ha sido cambiada exitosamente! Ya puedes iniciar sesión.'
                )
                
                # Respuesta AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Contraseña cambiada exitosamente',
                        'redirect_url': '/auth/login/'
                    })
                
                return redirect('autenticacion:login')
            else:
                messages.error(request, 'Código inválido o expirado.')
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Código inválido o expirado'
                    })
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for error in form.errors.values():
                messages.error(request, error[0])
    else:
        form = VerificarCodigoCambioForm()
    
    context = {
        'form': form,
        'usuario': usuario,
        'email': usuario.email,
        'is_recovery': True,  # Flag para indicar que es recuperación (no cambio logueado)
    }
    
    return render(request, 'autenticacion/verificar_codigo_recuperacion.html', context)


@csrf_protect
def resetear_password(request, token):
    """Vista para resetear contraseña con token"""
    if request.user.is_authenticated:
        return redirect('autenticacion:dashboard')
    
    # Validar token
    token_reset = validar_token_reset(token)
    if not token_reset:
        messages.error(request, 'El enlace de recuperación es inválido o ha expirado.')
        return redirect('autenticacion:recuperar_password')
    
    if request.method == 'POST':
        form = ResetearPasswordForm(request.POST)
        if form.is_valid():
            # Cambiar contraseña
            usuario = token_reset.usuario
            nueva_password = form.cleaned_data['new_password1']
            usuario.set_password(nueva_password)
            usuario.save()
            
            # Marcar token como usado
            marcar_token_usado(token_reset)
            
            messages.success(
                request, 
                'Tu contraseña ha sido restablecida exitosamente. '
                'Ya puedes iniciar sesión con tu nueva contraseña.'
            )
            
            # Respuesta AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Contraseña restablecida exitosamente'
                })
            
            return redirect('autenticacion:login')
        else:
            # Respuesta AJAX con errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            
            for error in form.errors.values():
                messages.error(request, error[0])
    else:
        form = ResetearPasswordForm()
    
    context = {
        'form': form,
        'token': token,
        'usuario': token_reset.usuario,
    }
    
    return render(request, 'autenticacion/resetear_password.html', context)


@login_required
@require_http_methods(["DELETE"])
def eliminar_avatar(request):
    """Vista AJAX para eliminar avatar del usuario"""
    try:
        usuario = request.user
        if usuario.avatar:
            # Eliminar archivo físico
            try:
                usuario.avatar.delete(save=False)
            except:
                pass  # Ignorar errores al eliminar archivo
            
            # Limpiar campo en BD
            usuario.avatar = None
            usuario.save(update_fields=['avatar'])
            
            return JsonResponse({
                'success': True,
                'message': 'Avatar eliminado exitosamente'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No tienes avatar para eliminar'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': 'Error eliminando avatar'
        })


@require_http_methods(["POST"])
def verificar_email_existente(request):
    """Vista API para verificar si un email existe en el sistema"""
    import json
    
    try:
        data = json.loads(request.body)
        email = data.get('email', '').strip()
        
        if not email:
            return JsonResponse({
                'existe': False,
                'message': 'Email requerido'
            })
        
        # Verificar si existe un usuario activo con ese email
        existe = Usuario.objects.filter(
            email__iexact=email,
            estado='ACTIVO'
        ).exists()
        
        return JsonResponse({
            'existe': existe,
            'message': 'Email registrado' if existe else 'Email no encontrado'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'existe': False,
            'message': 'Datos inválidos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'existe': False,
            'message': 'Error verificando email'
        }, status=500)


# ==================== GESTIÓN DE USUARIOS (CRUD) ====================

from django.core.paginator import Paginator
from django.db.models import Q, Count
from .decorators import admin_only, permission_required, login_required_custom, estado_usuario_activo
from .models import Rol, Sesion


@login_required
@permission_required('usuarios.ver')
def usuario_listar(request):
    """Lista de usuarios con búsqueda y filtros"""
    query = request.GET.get('q', '')
    estado_filter = request.GET.get('estado', '')
    rol_filter = request.GET.get('rol', '')
    per_page = request.GET.get('per_page', '10')
    
    # Consulta base
    usuarios = Usuario.objects.select_related('rol').all()
    
    # Aplicar filtros
    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(email__icontains=query)
        )
    
    if estado_filter:
        usuarios = usuarios.filter(estado=estado_filter)
    
    if rol_filter:
        usuarios = usuarios.filter(rol_id=rol_filter)
    
    # Ordenar
    ordenar = request.GET.get('orden', '-created_at')
    usuarios = usuarios.order_by(ordenar)
    
    # Paginación
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 15, 20, 25, 30, 50, 100, 500, 1000, 10000]:
            per_page = 10
    except:
        per_page = 10
    
    paginator = Paginator(usuarios, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'usuarios': page_obj,
        'roles': Rol.objects.all(),
        'estados': Usuario.ESTADO_CHOICES,
        'query': query,
        'estado_filter': estado_filter,
        'rol_filter': int(rol_filter) if rol_filter else '',
        'per_page': per_page,
        'per_page_options': [5, 10, 15, 20, 25, 30, 50, 100, 500, 1000, 10000],
        'ordenar': ordenar,
        'total_usuarios': usuarios.count(),
    }
    
    return render(request, 'autenticacion/usuario_listar.html', context)


@login_required
@permission_required('usuarios.crear')
def usuario_crear(request):
    """Crear nuevo usuario"""
    if request.method == 'POST':
        errores = {}
        datos = {}
        
        try:
            # Capturar datos
            datos = {
                'username': request.POST.get('username', '').strip(),
                'nombres': request.POST.get('nombres', '').strip(),
                'apellidos': request.POST.get('apellidos', '').strip(),
                'email': request.POST.get('email', '').strip(),
                'telefono': request.POST.get('telefono', '').strip(),
                'area_unidad': request.POST.get('area_unidad', '').strip(),
                'observaciones': request.POST.get('observaciones', '').strip(),
                'rol_id': request.POST.get('rol_id', ''),
                'estado': request.POST.get('estado', 'ACTIVO'),
                'password': request.POST.get('password', ''),
                'password_confirm': request.POST.get('password_confirm', ''),
            }
            
            # Validaciones
            if not datos['username']:
                errores['username'] = 'El nombre de usuario es obligatorio'
            elif len(datos['username']) < 3:
                errores['username'] = 'El nombre de usuario debe tener al menos 3 caracteres'
            elif len(datos['username']) > 8:
                errores['username'] = 'El nombre de usuario no puede tener más de 8 caracteres'
            elif Usuario.objects.filter(username=datos['username']).exists():
                errores['username'] = 'Ya existe un usuario con este nombre'
                
            if not datos['nombres']:
                errores['nombres'] = 'Los nombres son obligatorios'
            elif len(datos['nombres']) < 2:
                errores['nombres'] = 'Los nombres deben tener al menos 2 caracteres'
            elif len(datos['nombres']) > 8:
                errores['nombres'] = 'Los nombres no pueden tener más de 8 caracteres'
                
            if not datos['apellidos']:
                errores['apellidos'] = 'Los apellidos son obligatorios'
            elif len(datos['apellidos']) < 2:
                errores['apellidos'] = 'Los apellidos deben tener al menos 2 caracteres'
            elif len(datos['apellidos']) > 8:
                errores['apellidos'] = 'Los apellidos no pueden tener más de 8 caracteres'
                
            if datos['email']:
                if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', datos['email']):
                    errores['email'] = 'Email inválido'
                elif Usuario.objects.filter(email=datos['email']).exists():
                    errores['email'] = 'Ya existe un usuario con este email'
                    
            if not datos['rol_id']:
                errores['rol_id'] = 'Debe seleccionar un rol'
            else:
                try:
                    rol = Rol.objects.get(id=datos['rol_id'])
                    datos['rol'] = rol
                except Rol.DoesNotExist:
                    errores['rol_id'] = 'El rol seleccionado no existe'
                    
            if not datos['password']:
                errores['password'] = 'La contraseña es obligatoria'
            elif len(datos['password']) < 6:
                errores['password'] = 'La contraseña debe tener al menos 6 caracteres'
            elif datos['password'] != datos['password_confirm']:
                errores['password_confirm'] = 'Las contraseñas no coinciden'
                
            if errores:
                messages.error(request, 'Corrija los errores en el formulario')
                context = {
                    'errores': errores,
                    'datos': datos,
                    'roles': Rol.objects.all(),
                    'estados': Usuario.ESTADO_CHOICES,
                }
                return render(request, 'autenticacion/usuario_crear.html', context)
            
            # Crear usuario
            usuario = Usuario.objects.create_user(
                username=datos['username'],
                password=datos['password'],
                nombres=datos['nombres'],
                apellidos=datos['apellidos'],
                email=datos['email'] or '',
                telefono=datos['telefono'] or None,
                area_unidad=datos['area_unidad'] or None,
                observaciones=datos['observaciones'] or None,
                rol=datos['rol'],
                estado=datos['estado']
            )
            
            messages.success(request, f'Usuario "{usuario.username}" creado exitosamente')
            return redirect('autenticacion:usuario_detalle', pk=usuario.pk)
            
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
            context = {
                'errores': errores,
                'datos': datos,
                'roles': Rol.objects.all(),
                'estados': Usuario.ESTADO_CHOICES,
            }
            return render(request, 'autenticacion/usuario_crear.html', context)
    
    # GET request
    context = {
        'roles': Rol.objects.all(),
        'estados': Usuario.ESTADO_CHOICES,
    }
    return render(request, 'autenticacion/usuario_crear.html', context)


@login_required
@permission_required('usuarios.ver')
def usuario_detalle(request, pk):
    """Detalle de usuario"""
    usuario = get_object_or_404(Usuario, pk=pk)
    
    # Estadísticas de sesiones recientes
    sesiones_recientes = Sesion.objects.filter(usuario=usuario).order_by('-created_at')[:10]
    total_sesiones = Sesion.objects.filter(usuario=usuario).count()
    
    # Verificar si puede editar
    puede_editar = request.user.rol and (
        request.user.rol.nombre == 'Administrador' or 
        request.user.pk == usuario.pk
    )
    
    context = {
        'usuario': usuario,
        'sesiones_recientes': sesiones_recientes,
        'total_sesiones': total_sesiones,
        'puede_editar': puede_editar,
    }
    return render(request, 'autenticacion/usuario_detalle.html', context)


@login_required_custom
@estado_usuario_activo
def usuario_editar(request, pk):
    """Editar usuario existente"""
    usuario = get_object_or_404(Usuario, pk=pk)
    
    # Solo administradores, superusuarios o el mismo usuario pueden editar
    if not (request.user.is_superuser or 
            (request.user.rol and request.user.rol.nombre == 'Administrador') or 
            request.user.pk == usuario.pk):
        messages.error(request, 'No tienes permisos para editar este usuario')
        return redirect('autenticacion:usuario_detalle', pk=pk)
    
    if request.method == 'POST':
        errores = {}
        datos = {}
        
        try:
            # Capturar datos
            print(f"=== DATOS POST RECIBIDOS ===")
            for key, value in request.POST.items():
                print(f"{key}: '{value}'")
            print(f"=== FIN DATOS POST ===")
            
            datos = {
                'username': request.POST.get('username', '').strip(),
                'nombres': request.POST.get('nombres', '').strip(),
                'apellidos': request.POST.get('apellidos', '').strip(),
                'email': request.POST.get('email', '').strip(),
                'telefono': request.POST.get('telefono', '').strip(),
                'direccion': request.POST.get('direccion', '').strip(),
                'rol_id': request.POST.get('rol_id', ''),
                'estado': request.POST.get('estado', 'ACTIVO'),
                'nueva_password': request.POST.get('nueva_password', '').strip(),
                'confirmar_password': request.POST.get('confirmar_password', '').strip(),
                'is_active': request.POST.get('is_active') == 'on',
                'is_staff': request.POST.get('is_staff') == 'on',
                'is_superuser': request.POST.get('is_superuser') == 'on',
            }
            
            print(f"=== DATOS PROCESADOS ===")
            for key, value in datos.items():
                print(f"{key}: '{value}'")
            
            # Validaciones
            print(f"=== VALIDACIONES ===")
            
            # Validar username
            if not datos['username']:
                errores['username'] = 'El nombre de usuario es obligatorio'
                print(f"ERROR: Username vacío")
            elif len(datos['username']) < 3:
                errores['username'] = 'El nombre de usuario debe tener al menos 3 caracteres'
                print(f"ERROR: Username muy corto")
            elif len(datos['username']) > 8:
                errores['username'] = 'El nombre de usuario no puede tener más de 8 caracteres'
                print(f"ERROR: Username muy largo")
            elif Usuario.objects.filter(username=datos['username']).exclude(pk=usuario.pk).exists():
                errores['username'] = 'Ya existe otro usuario con este nombre de usuario'
                print(f"ERROR: Username duplicado")
                
            if not datos['nombres']:
                errores['nombres'] = 'Los nombres son obligatorios'
                print(f"ERROR: Nombres vacío")
            elif len(datos['nombres']) < 2:
                errores['nombres'] = 'Los nombres deben tener al menos 2 caracteres'
                print(f"ERROR: Nombres muy cortos")
            elif len(datos['nombres']) > 8:
                errores['nombres'] = 'Los nombres no pueden tener más de 8 caracteres'
                print(f"ERROR: Nombres muy largos")
                
            if not datos['apellidos']:
                errores['apellidos'] = 'Los apellidos son obligatorios'
                print(f"ERROR: Apellidos vacío")
            elif len(datos['apellidos']) < 2:
                errores['apellidos'] = 'Los apellidos deben tener al menos 2 caracteres'
                print(f"ERROR: Apellidos muy cortos")
            elif len(datos['apellidos']) > 8:
                errores['apellidos'] = 'Los apellidos no pueden tener más de 8 caracteres'
                print(f"ERROR: Apellidos muy largos")
                
            if not datos['email']:
                errores['email'] = 'El email es obligatorio'
                print(f"ERROR: Email vacío")
            elif not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', datos['email']):
                errores['email'] = 'Email inválido'
                print(f"ERROR: Email formato inválido")
            elif Usuario.objects.filter(email=datos['email']).exclude(pk=usuario.pk).exists():
                errores['email'] = 'Ya existe otro usuario con este email'
                print(f"ERROR: Email duplicado")
                
            # Validar contraseñas si se proporcionaron
            if datos['nueva_password'] or datos['confirmar_password']:
                if datos['nueva_password'] != datos['confirmar_password']:
                    errores['confirmar_password'] = 'Las contraseñas no coinciden'
                    print(f"ERROR: Contraseñas no coinciden")
                elif len(datos['nueva_password']) < 8:
                    errores['nueva_password'] = 'La contraseña debe tener al menos 8 caracteres'
                    print(f"ERROR: Contraseña muy corta")
                    
            # Solo administradores pueden cambiar rol y estado
            if request.user.is_superuser or (request.user.rol and request.user.rol.nombre == 'Administrador'):
                if not datos['rol_id']:
                    errores['rol_id'] = 'Debe seleccionar un rol'
                else:
                    try:
                        rol = Rol.objects.get(id=datos['rol_id'])
                        datos['rol'] = rol
                    except Rol.DoesNotExist:
                        errores['rol_id'] = 'El rol seleccionado no existe'
            else:
                # Mantener rol y estado actuales si no es admin
                datos['rol'] = usuario.rol
                datos['estado'] = usuario.estado
                
            if errores:
                print(f"=== ERRORES ENCONTRADOS ===")
                for campo, error in errores.items():
                    print(f"{campo}: {error}")
                print(f"=== FIN ERRORES ===")
                
                messages.error(request, 'Corrija los errores en el formulario')
                
                # Si es petición AJAX, devolver JSON con errores
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Corrija los errores en el formulario',
                        'errors': errores
                    })
                
                context = {
                    'usuario': usuario,
                    'errores': errores,
                    'datos': datos,
                    'roles': Rol.objects.all(),
                    'estados': Usuario.ESTADO_CHOICES,
                    'es_admin': request.user.is_superuser or (request.user.rol and request.user.rol.nombre == 'Administrador'),
                }
                return render(request, 'autenticacion/usuario_editar.html', context)
            
            # Actualizar usuario
            print(f"=== ACTUALIZANDO USUARIO ===")
            print(f"Username: '{datos['username']}' -> usuario.username antes: '{usuario.username}'")
            print(f"Nombres: '{datos['nombres']}' -> usuario.nombres antes: '{usuario.nombres}'")
            print(f"Apellidos: '{datos['apellidos']}' -> usuario.apellidos antes: '{usuario.apellidos}'")
            print(f"Email: '{datos['email']}' -> usuario.email antes: '{usuario.email}'")
            
            # Actualizar campos básicos
            usuario.username = datos['username']
            usuario.nombres = datos['nombres']
            usuario.apellidos = datos['apellidos'] 
            usuario.email = datos['email']
            usuario.telefono = datos['telefono'] or None
            usuario.direccion = datos['direccion'] or None
            usuario.is_active = datos['is_active']
            
            # Cambiar contraseña si se proporcionó
            if datos['nueva_password']:
                usuario.set_password(datos['nueva_password'])
                print(f"Contraseña actualizada")
            
            print(f"Después de asignar - Username: '{usuario.username}', Nombres: '{usuario.nombres}', Apellidos: '{usuario.apellidos}'")
            
            # Solo administradores pueden cambiar rol, estado y permisos especiales
            if request.user.is_superuser or (request.user.rol and request.user.rol.nombre == 'Administrador'):
                usuario.rol = datos['rol']
                usuario.estado = datos['estado']
                
                # Solo superusuarios pueden cambiar is_staff e is_superuser
                if request.user.is_superuser:
                    usuario.is_staff = datos['is_staff']
                    usuario.is_superuser = datos['is_superuser']
                    
                print(f"Admin - Rol: {usuario.rol}, Estado: {usuario.estado}, is_staff: {usuario.is_staff}, is_superuser: {usuario.is_superuser}")
            
            # Procesar avatar si se subió uno nuevo
            if 'avatar' in request.FILES:
                from .utils import procesar_avatar
                avatar_procesado = procesar_avatar(request.FILES['avatar'], usuario)
                if avatar_procesado:
                    usuario.avatar = avatar_procesado
                    print(f"Avatar actualizado: {usuario.avatar}")
            
            usuario.save()
            print(f"Usuario guardado - ID: {usuario.pk}")
            
            # Recargar desde DB para verificar
            usuario.refresh_from_db()
            print(f"Después de refresh - Nombres: '{usuario.nombres}', Apellidos: '{usuario.apellidos}'")
            
            messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente')
            
            # Si es petición AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Usuario "{usuario.username}" actualizado exitosamente'
                })
            
            return redirect('autenticacion:usuario_detalle', pk=usuario.pk)
            
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
            
            # Si es petición AJAX, devolver JSON con error
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar usuario: {str(e)}'
                })
    
    # GET request
    context = {
        'usuario': usuario,
        'roles': Rol.objects.all(),
        'estados': Usuario.ESTADO_CHOICES,
        'es_admin': request.user.is_superuser or (request.user.rol and request.user.rol.nombre == 'Administrador'),
        'es_superuser': request.user.is_superuser,
    }
    return render(request, 'autenticacion/usuario_editar.html', context)


@login_required
@permission_required('usuarios.editar')
@require_http_methods(["POST"])
def usuario_cambiar_estado(request, pk):
    """Cambiar estado de usuario (AJAX)"""
    usuario = get_object_or_404(Usuario, pk=pk)
    nuevo_estado = request.POST.get('estado')
    
    if nuevo_estado not in dict(Usuario.ESTADO_CHOICES):
        return JsonResponse({
            'success': False,
            'message': 'Estado inválido'
        })
    
    try:
        usuario.estado = nuevo_estado
        usuario.save(update_fields=['estado'])
        
        return JsonResponse({
            'success': True,
            'message': f'Estado cambiado a {usuario.get_estado_display()}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al cambiar estado: {str(e)}'
        })


@login_required
@permission_required('usuarios.editar')
@require_http_methods(["POST"])
def usuario_resetear_password(request, pk):
    """Resetear contraseña de usuario"""
    usuario = get_object_or_404(Usuario, pk=pk)
    nueva_password = request.POST.get('nueva_password', '')
    
    if len(nueva_password) < 6:
        return JsonResponse({
            'success': False,
            'message': 'La contraseña debe tener al menos 6 caracteres'
        })
    
    try:
        usuario.set_password(nueva_password)
        usuario.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Contraseña restablecida exitosamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al restablecer contraseña: {str(e)}'
        })


@login_required
@permission_required('usuarios.ver')
def usuario_historial(request, pk):
    """Historial de accesos del usuario"""
    usuario = get_object_or_404(Usuario, pk=pk)
    
    # Sesiones con paginación
    sesiones = Sesion.objects.filter(usuario=usuario).order_by('-created_at')
    paginator = Paginator(sesiones, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'usuario': usuario,
        'sesiones': page_obj,
    }
    return render(request, 'autenticacion/usuario_historial.html', context)


@login_required
@admin_only
@csrf_protect
def usuario_eliminar(request, pk):
    """Eliminar usuario (solo admins)"""
    usuario = get_object_or_404(Usuario, pk=pk)
    
    # No permitir eliminar superusers o el mismo usuario
    if usuario.is_superuser or usuario.pk == request.user.pk:
        messages.error(request, 'No se puede eliminar este usuario')
        return redirect('autenticacion:usuario_detalle', pk=pk)
    
    if request.method == 'POST':
        nombre_usuario = usuario.get_full_name()
        try:
            usuario.delete()
            messages.success(request, f'Usuario "{nombre_usuario}" eliminado exitosamente')
            return redirect('autenticacion:usuario_listar')
        except Exception as e:
            messages.error(request, f'Error al eliminar usuario: {str(e)}')
            return redirect('autenticacion:usuario_detalle', pk=pk)
    
    context = {
        'usuario': usuario,
    }
    return render(request, 'autenticacion/usuario_eliminar.html', context)


# ==================== GESTIÓN DE ROLES ====================

@login_required
@admin_only
def rol_listar(request):
    """Lista de roles"""
    roles = Rol.objects.annotate(
        total_usuarios=Count('usuarios')
    ).order_by('nombre')
    
    context = {
        'roles': roles,
    }
    return render(request, 'autenticacion/rol_listar.html', context)


@login_required
@admin_only
def rol_crear(request):
    """Crear nuevo rol"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        permisos_json = request.POST.get('permisos', '{}')
        
        errores = {}
        
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio'
        elif Rol.objects.filter(nombre=nombre).exists():
            errores['nombre'] = 'Ya existe un rol con este nombre'
            
        try:
            import json
            permisos = json.loads(permisos_json)
        except:
            errores['permisos'] = 'Los permisos deben ser un JSON válido'
            permisos = {}
            
        if errores:
            messages.error(request, 'Corrija los errores en el formulario')
            context = {
                'errores': errores,
                'nombre': nombre,
                'descripcion': descripcion,
                'permisos_json': permisos_json,
            }
            return render(request, 'autenticacion/rol_crear.html', context)
        
        try:
            rol = Rol.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                permisos=permisos
            )
            messages.success(request, f'Rol "{rol.nombre}" creado exitosamente')
            return redirect('autenticacion:rol_listar')
        except Exception as e:
            messages.error(request, f'Error al crear rol: {str(e)}')
    
    return render(request, 'autenticacion/rol_crear.html')


@login_required
@admin_only
def rol_editar(request, pk):
    """Editar rol existente"""
    rol = get_object_or_404(Rol, pk=pk)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        permisos_json = request.POST.get('permisos', '{}')
        
        errores = {}
        
        if not nombre:
            errores['nombre'] = 'El nombre es obligatorio'
        elif Rol.objects.filter(nombre=nombre).exclude(pk=rol.pk).exists():
            errores['nombre'] = 'Ya existe otro rol con este nombre'
            
        try:
            import json
            permisos = json.loads(permisos_json)
        except:
            errores['permisos'] = 'Los permisos deben ser un JSON válido'
            permisos = rol.permisos or {}
            
        if errores:
            messages.error(request, 'Corrija los errores en el formulario')
            context = {
                'rol': rol,
                'errores': errores,
                'nombre': nombre,
                'descripcion': descripcion,
                'permisos_json': permisos_json,
            }
            return render(request, 'autenticacion/rol_editar.html', context)
        
        try:
            rol.nombre = nombre
            rol.descripcion = descripcion
            rol.permisos = permisos
            rol.save()
            
            messages.success(request, f'Rol "{rol.nombre}" actualizado exitosamente')
            return redirect('autenticacion:rol_listar')
        except Exception as e:
            messages.error(request, f'Error al actualizar rol: {str(e)}')
    
    context = {
        'rol': rol,
        'permisos_json': json.dumps(rol.permisos or {}, indent=2),
    }
    return render(request, 'autenticacion/rol_editar.html', context)


# ==================== EXPORTACIÓN A EXCEL ====================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.http import HttpResponse


def crear_estilos_excel():
    """Crear estilos reutilizables para Excel"""
    return {
        'header': {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'data': {
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        }
    }


@login_required
def export_usuarios_excel(request):
    """Exportar usuarios a Excel - Solo administradores"""
    # Verificar que sea administrador
    if not request.user.is_superuser and (not hasattr(request.user, 'rol') or not request.user.rol or request.user.rol.nombre != 'ADMIN'):
        messages.error(request, 'No tienes permisos para exportar usuarios.')
        return redirect('autenticacion:usuario_listar')
    
    try:
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Estilos
        estilos = crear_estilos_excel()
        
        # Encabezados
        headers = [
            'ID', 'Usuario', 'Nombres', 'Apellidos', 'Email', 'Teléfono',
            'Rol', 'Estado', 'Es Superusuario', 'Último Acceso',
            'Fecha Registro', 'Última Modificación'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = estilos['header']['font']
            cell.alignment = estilos['header']['alignment']
            cell.fill = estilos['header']['fill']
            cell.border = estilos['header']['border']
        
        # Obtener datos
        usuarios = Usuario.objects.select_related('rol').order_by('username')
        
        # Agregar datos
        for row_num, usuario in enumerate(usuarios, 2):
            data = [
                usuario.id,
                usuario.username,
                usuario.nombres,  # Usar 'nombres' en lugar de 'first_name'
                usuario.apellidos,  # Usar 'apellidos' en lugar de 'last_name'
                usuario.email,
                usuario.telefono or 'Sin teléfono',
                usuario.rol.nombre if usuario.rol else 'Sin rol',
                usuario.get_estado_display(),
                'Sí' if usuario.is_superuser else 'No',
                usuario.ultimo_acceso.strftime('%d/%m/%Y %H:%M') if usuario.ultimo_acceso else 'Nunca',
                usuario.date_joined.strftime('%d/%m/%Y %H:%M') if usuario.date_joined else '',
                usuario.updated_at.strftime('%d/%m/%Y %H:%M') if usuario.updated_at else ''  # Usar 'updated_at' en lugar de 'fecha_modificacion'
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = estilos['data']['alignment']
                cell.border = estilos['data']['border']
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar usuarios: {str(e)}')
        return redirect('autenticacion:usuario_listar')
